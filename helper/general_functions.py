import os
import csv
import ast
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
import json
import openpyxl
from helper.utils import clean_text


def split_text(text, max_length=300):
    text = clean_text(text)
    sentences = text.split('. ')
    chunks = []
    current_chunk = []
    try:
        for sentence in sentences:
            if len(' '.join(current_chunk + [sentence])) > max_length:
                chunks.append(' '.join(current_chunk))
                current_chunk = [sentence]
            else:
                current_chunk.append(sentence)
        if current_chunk:
            chunks.append(' '.join(current_chunk))
    except:
        print("split_text", text)
    return chunks

def create_and_write_csv(file_name, data, method_name):   
    dictory_path = "model/DeepCGSR/feature"
    filename = os.path.join(dictory_path, file_name + '.csv')
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Key', 'Array'])
        for key, value in data.items():
            if isinstance(value, np.ndarray):
                array_str = str(value.tolist())
            else:
                array_str = str(value)
            writer.writerow([key, array_str])

def load_data_from_csv(file_path):
    data = {}
    with open(file_path, 'r') as csv_file:
        csv_reader = csv.DictReader(csv_file)
        for row in csv_reader:
            key = row['Key']
            array_str = row['Array']
            if array_str.startswith('[') and array_str.endswith(']'):
                data[key] = np.array(eval(array_str))
            else:
                array = np.array(ast.literal_eval(array_str))
                data[key] = array
    return data

def read_csv_file(csv_file):
    keys = []
    values = []

    with open(csv_file, 'r', encoding='utf-8') as file:
        csv_reader = csv.reader(file)
        next(csv_reader)  # Bỏ qua header
        for row in csv_reader:
            key = row[0]  # ID là cột đầu tiên
            value_str = row[1]  # Giá trị là cột thứ hai
            value = ast.literal_eval(value_str)  # Chuyển đổi chuỗi thành vector
            keys.append(key)
            values.append(np.array(value))

    return keys, values

def read_and_split_dataset(file_path, train_ratio=0.7, valid_ratio=0.15, test_ratio=0.15, random_state=None):
    assert train_ratio + valid_ratio + test_ratio == 1.0, "Train, validation and test ratios must sum to 1.0"

    with open(file_path, 'r') as file:
        data = json.load(file)
    df = pd.DataFrame(data)
    train_df, remaining_df = train_test_split(df, train_size=train_ratio, random_state=random_state)
    remaining_ratio = valid_ratio / (valid_ratio + test_ratio)
    valid_df, test_df = train_test_split(remaining_df, train_size=remaining_ratio, random_state=random_state)
    return train_df, valid_df, test_df

def read_and_split_csv_dataset(file_path, train_ratio=0.7, valid_ratio=0.15, test_ratio=0.15, random_state=None):

    assert train_ratio + valid_ratio + test_ratio == 1.0, "Train, validation and test ratios must sum to 1.0"
    df = pd.read_csv(file_path)
    train_df, remaining_df = train_test_split(df, train_size=train_ratio, random_state=random_state)
    remaining_ratio = valid_ratio / (valid_ratio + test_ratio)
    valid_df, test_df = train_test_split(remaining_df, train_size=remaining_ratio, random_state=random_state)
    return train_df, valid_df, test_df

def save_to_excel(values, headers, output_path):
    if os.path.exists(output_path):
        workbook = openpyxl.load_workbook(output_path)
        sheet = workbook.active
        start_row = sheet.max_row
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        start_row = 1 

        for index, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=index, value=header)
            
    for row_index, row_values in enumerate(values, start=start_row + 1):
        for col_index, value in enumerate(row_values, start=1):
            sheet.cell(row=row_index, column=col_index, value=round(value, 4))

    workbook.save(output_path)
    print(f"Đã lưu danh sách giá trị vào tệp '{output_path}' thành công.")